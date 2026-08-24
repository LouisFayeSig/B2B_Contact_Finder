from __future__ import annotations

import logging
import unittest
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from threading import Barrier

from openpyxl import Workbook

from app.batch_processor import BatchProcessor
from app.config import AppConfig
from app.excel_service import ExcelService
from app.models import CompanyResult, CompanyRow, ProcessingStats, ProcessingStatus


class SuccessfulSearch:
    def __init__(self, result: CompanyResult) -> None:
        self.result = result

    def search_company_contact(self, company: CompanyRow) -> CompanyResult:
        return self.result


class FailingSearch:
    def search_company_contact(self, company: CompanyRow) -> CompanyResult:
        raise RuntimeError("incident temporaire")


class SynchronizedSearch:
    def __init__(self, expected_workers: int) -> None:
        self.barrier = Barrier(expected_workers, timeout=2)

    def search_company_contact(self, company: CompanyRow) -> CompanyResult:
        self.barrier.wait()
        return CompanyResult(
            email=f"row{company.row_index}@example.com",
            identity_verified=True,
            identity_match_type="siret",
        )


def make_config(*, overwrite_existing: bool = False) -> AppConfig:
    return AppConfig(
        input_excel_path=Path("unused.xlsx"),
        sheet_name=None,
        start_row=2,
        max_rows=None,
        batch_size=100,
        save_every_batch=True,
        skip_if_filled=True,
        overwrite_existing=overwrite_existing,
        azure_foundry_endpoint="https://example.openai.azure.com/openai/v1/",
        azure_foundry_api_key="test",
        azure_foundry_model_deployment="gpt-5.6-luna",
        azure_foundry_reasoning_effort="none",
        web_search_context_size="default",
        search_audit_enabled=False,
        max_workers=2,
        request_timeout=90,
        max_retries=3,
        retry_wait_seconds=0,
        sleep_between_calls=0,
        sleep_between_batches=0,
        log_level="INFO",
        create_backup=False,
        azure_foundry_ca_bundle=None,
        log_file_path=Path("unused.log"),
        journal_file_path=Path("unused.pending.jsonl"),
    )


class BatchProcessorTests(unittest.TestCase):
    def setUp(self) -> None:
        self.logger = logging.getLogger("tests.batch")
        self.service = ExcelService(Path("unused.xlsx"), None, self.logger)
        self.service._workbook = Workbook()
        self.service._sheet = self.service.workbook.active

    def test_technical_error_preserves_contacts_and_remains_retryable(self) -> None:
        self.service.sheet.cell(row=2, column=self.service.columns.email, value="existing@example.com")
        processor = BatchProcessor(make_config(), self.service, FailingSearch(), self.logger)
        stats = ProcessingStats()

        processor._process_single_company(self._company(2), stats)

        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.email).value,
            "existing@example.com",
        )
        self.assertEqual(self.service.read_status(2), ProcessingStatus.TECHNICAL_ERROR)
        self.assertFalse(processor._should_skip_row(2))
        self.assertEqual(stats.technical_error_rows, 1)

    def test_partial_row_is_completed_without_overwriting_existing_email(self) -> None:
        self.service.sheet.cell(row=2, column=self.service.columns.email, value="existing@example.com")
        search = SuccessfulSearch(
            CompanyResult(
                email="new@example.com",
                phone="01 23 45 67 89",
                website="example.com",
                identity_verified=True,
                identity_match_type="siret",
            )
        )
        processor = BatchProcessor(make_config(), self.service, search, self.logger)

        processor._process_single_company(self._company(2), ProcessingStats())

        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.email).value,
            "existing@example.com",
        )
        self.assertEqual(self.service.read_status(2), ProcessingStatus.SUCCESS)
        self.assertTrue(processor._should_skip_row(2))

    def test_not_found_is_terminal_but_distinct_from_technical_error(self) -> None:
        processor = BatchProcessor(
            make_config(),
            self.service,
            SuccessfulSearch(CompanyResult.not_found()),
            self.logger,
        )
        stats = ProcessingStats()

        processor._process_single_company(self._company(2), stats)

        self.assertEqual(self.service.read_status(2), ProcessingStatus.NOT_FOUND)
        self.assertTrue(processor._should_skip_row(2))
        self.assertEqual(stats.not_found_rows, 1)

    def test_invalid_input_is_reprocessed_after_source_data_is_fixed(self) -> None:
        processor = BatchProcessor(
            make_config(),
            self.service,
            SuccessfulSearch(CompanyResult.not_found()),
            self.logger,
        )
        invalid_company = CompanyRow(row_index=2, siret="", company_name="Entreprise Test")

        processor._process_single_company(invalid_company, ProcessingStats())
        self.assertEqual(self.service.read_status(2), ProcessingStatus.INVALID_INPUT)
        self.assertTrue(processor._should_skip_row(2))

        self.service.sheet.cell(row=2, column=self.service.columns.siret, value="12345678901234")
        self.service.sheet.cell(row=2, column=self.service.columns.company_name, value="Entreprise Test")

        self.assertFalse(processor._should_skip_row(2))

    def test_batch_searches_run_concurrently_and_excel_writes_are_committed(self) -> None:
        for row_index in (2, 3):
            self.service.sheet.cell(
                row=row_index,
                column=self.service.columns.siret,
                value=f"1234567890123{row_index}",
            )
            self.service.sheet.cell(
                row=row_index,
                column=self.service.columns.company_name,
                value=f"Entreprise {row_index}",
            )
        processor = BatchProcessor(
            make_config(),
            self.service,
            SynchronizedSearch(expected_workers=2),
            self.logger,
        )
        stats = ProcessingStats()

        with ThreadPoolExecutor(max_workers=2) as executor:
            processor._process_batch([2, 3], stats, executor)

        self.assertEqual(stats.success_rows, 2)
        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.email).value,
            "row2@example.com",
        )
        self.assertEqual(self.service.read_status(3), ProcessingStatus.SUCCESS)

    def _company(self, row_index: int) -> CompanyRow:
        return CompanyRow(
            row_index=row_index,
            siret="12345678901234",
            company_name="Entreprise Test",
        )


if __name__ == "__main__":
    unittest.main()
