from __future__ import annotations

import logging
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.excel_service import ExcelService
from app.journal import ProcessingJournal
from app.models import CompanyResult, CompanyRow, ProcessingStatus


class ProcessingJournalTests(unittest.TestCase):
    def test_pending_result_is_replayed_and_journal_is_cleared(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            excel_path = root / "companies.xlsx"
            journal_path = root / "pending.jsonl"
            service = ExcelService(
                excel_path,
                None,
                logging.getLogger("tests.journal"),
                audit_enabled=True,
            )
            service._workbook = Workbook()
            service._sheet = service.workbook.active
            service.sheet.cell(row=2, column=service.columns.siret, value="12345678901234")
            service.sheet.cell(row=2, column=service.columns.company_name, value="Entreprise Test")
            company = CompanyRow(
                row_index=2,
                siret="12345678901234",
                company_name="Entreprise Test",
            )
            result = CompanyResult(
                email="contact@example.com",
                email_source="https://example.com/contact",
                identity_verified=True,
                identity_match_type="siret",
                identity_source="https://example.com/legal",
            )
            journal = ProcessingJournal(journal_path, logging.getLogger("tests.journal"))

            journal.append(
                company,
                result,
                ProcessingStatus.SUCCESS,
                overwrite_existing=False,
            )
            recovered = journal.recover(service)

            self.assertEqual(recovered, 1)
            self.assertEqual(
                service.sheet.cell(row=2, column=service.columns.email).value,
                "contact@example.com",
            )
            self.assertEqual(service.read_status(2), ProcessingStatus.SUCCESS)
            self.assertFalse(journal_path.exists())


if __name__ == "__main__":
    unittest.main()
