from __future__ import annotations

import logging
import unittest
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook

from app.excel_columns import resolve_excel_columns
from app.excel_service import ExcelService
from app.models import NOT_FOUND_LABEL, CompanyResult, ProcessingStatus, WebsiteType


class ExcelServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.service = ExcelService(
            Path("unused.xlsx"),
            None,
            logging.getLogger("tests.excel"),
            audit_enabled=True,
        )
        self.service._workbook = Workbook()
        self.service._sheet = self.service.workbook.active

    def test_partial_existing_values_are_not_overwritten(self) -> None:
        self.service.sheet.cell(row=2, column=self.service.columns.email, value="existing@example.com")

        self.service.write_result(
            2,
            CompanyResult(
                email="new@example.com",
                phone="01 23 45 67 89",
                website="example.com",
                identity_verified=True,
                identity_match_type="siret",
            ),
        )

        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.email).value,
            "existing@example.com",
        )
        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.phone).value,
            "01 23 45 67 89",
        )

    def test_overwrite_mode_replaces_existing_values(self) -> None:
        self.service.sheet.cell(row=2, column=self.service.columns.email, value="existing@example.com")

        self.service.write_result(
            2,
            CompanyResult(
                email="new@example.com",
                identity_verified=True,
                identity_match_type="siret",
            ),
            overwrite_existing=True,
        )

        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.email).value,
            "new@example.com",
        )

    def test_not_found_placeholder_can_be_replaced_without_overwrite(self) -> None:
        self.service.sheet.cell(
            row=2,
            column=self.service.columns.email,
            value=NOT_FOUND_LABEL,
        )

        self.service.write_result(
            2,
            CompanyResult(
                email="new@example.com",
                identity_verified=True,
                identity_match_type="siret",
            ),
        )

        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.email).value,
            "new@example.com",
        )

    def test_sources_and_status_round_trip(self) -> None:
        self.service.write_result(
            2,
            CompanyResult(sources=["https://example.com/contact"]),
        )
        self.service.write_status(2, ProcessingStatus.NOT_FOUND)

        self.assertEqual(self.service.read_sources(2), ["https://example.com/contact"])
        self.assertEqual(self.service.read_status(2), ProcessingStatus.NOT_FOUND)
        self.assertFalse(self.service.has_any_contact(2))
        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.email).value,
            NOT_FOUND_LABEL,
        )

    def test_result_headers_use_first_columns_after_business_data(self) -> None:
        self.service.sheet.cell(row=1, column=19, value="Contact")
        self.service.sheet.cell(row=1, column=25, value="Commentaire")

        self.service._ensure_result_headers()

        self.assertEqual(self.service.sheet.cell(row=1, column=19).value, "Contact")
        self.assertEqual(self.service.sheet.cell(row=1, column=25).value, "Commentaire")
        self.assertEqual(self.service.sheet.cell(row=1, column=26).value, "Enrichment Status")
        self.assertEqual(self.service.sheet.cell(row=1, column=27).value, "Enrichment Sources")
        self.assertEqual(self.service.sheet.cell(row=1, column=47).value, "Site Web Type")

    def test_website_type_is_written_without_changing_selected_website(self) -> None:
        result = CompanyResult(
            website="https://www.societe.com/entreprise-test",
            website_type="directory",
            website_source="https://www.societe.com/entreprise-test",
            identity_verified=True,
            identity_match_type="siret",
            identity_source="https://www.societe.com/entreprise-test",
        )

        self.service.write_result(2, result)

        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.website).value,
            "https://www.societe.com/entreprise-test",
        )
        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.website_type).value,
            WebsiteType.DIRECTORY.value,
        )

    def test_sources_are_not_written_when_audit_is_disabled(self) -> None:
        self.service.audit_enabled = False

        self.service.write_result(
            2,
            CompanyResult(sources=["https://example.com/contact"]),
        )
        self.service._ensure_result_headers()

        self.assertEqual(self.service.read_sources(2), [])
        self.assertIsNone(self.service.sheet.cell(row=1, column=27).value)

    def test_field_sources_and_search_metadata_are_written(self) -> None:
        result = CompanyResult(
            email="contact@example.com",
            email_source="https://example.com/contact",
            sources=["https://example.com/contact"],
            identity_verified=True,
            identity_match_type="siret",
            identity_source="https://example.com/legal",
            searched_at_utc=datetime(2026, 8, 24, 12, 0, tzinfo=UTC),
            model_deployment="gpt-5.6-luna",
            model_snapshot="gpt-5.6-luna-2026-07-09",
            response_id="resp_123",
            input_tokens=120,
            output_tokens=45,
            total_tokens=165,
            web_search_calls=1,
        )

        self.service.write_result(2, result)

        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.email_source).value,
            "https://example.com/contact",
        )
        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.identity_match_type).value,
            "siret",
        )
        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.response_id).value,
            "resp_123",
        )
        self.assertEqual(
            self.service.sheet.cell(row=2, column=self.service.columns.total_tokens).value,
            165,
        )

    def test_read_company_row_uses_resolved_columns(self) -> None:
        headers = {
            1: "VILLE",
            2: "raison_sociale",
            4: "siret",
            6: "CODE POSTAL",
            8: "Adresse",
        }
        for column_index, header in headers.items():
            self.service.sheet.cell(row=1, column=column_index, value=header)
        self.service.sheet.cell(row=2, column=1, value="Nantes")
        self.service.sheet.cell(row=2, column=2, value="Entreprise Test")
        self.service.sheet.cell(row=2, column=4, value="12345678901234")
        self.service.sheet.cell(row=2, column=6, value=44000)
        self.service.sheet.cell(row=2, column=8, value="1 rue du Test")
        self.service.columns = resolve_excel_columns(self.service.sheet)

        company = self.service.read_company_row(2)

        self.assertEqual(company.siret, "12345678901234")
        self.assertEqual(company.company_name, "Entreprise Test")
        self.assertEqual(company.address, "1 rue du Test")
        self.assertEqual(company.postal_code, "44000")
        self.assertEqual(company.city, "Nantes")


if __name__ == "__main__":
    unittest.main()
