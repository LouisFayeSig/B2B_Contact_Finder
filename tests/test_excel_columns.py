from __future__ import annotations

import unittest

from openpyxl import Workbook

from app.excel_columns import MissingExcelColumnError, normalize_excel_header, resolve_excel_columns


class ExcelColumnResolutionTests(unittest.TestCase):
    def test_header_normalization_ignores_case_accents_and_separators(self) -> None:
        self.assertEqual(normalize_excel_header("  DÉNOMINATION_sociale- "), "denomination sociale")

    def test_reordered_v2_headers_are_resolved_and_results_are_appended(self) -> None:
        sheet = Workbook().active
        headers = {
            3: "SIRET",
            5: "Raison sociale",
            6: "ADRESSE",
            8: "Code postal",
            9: "VILLE",
        }
        for column_index, header in headers.items():
            sheet.cell(row=1, column=column_index, value=header)

        columns = resolve_excel_columns(sheet)

        self.assertEqual(columns.siret, 3)
        self.assertEqual(columns.company_name, 5)
        self.assertEqual(columns.address, 6)
        self.assertEqual(columns.postal_code, 8)
        self.assertEqual(columns.city, 9)
        self.assertEqual(columns.email, 10)
        self.assertEqual(columns.phone, 11)
        self.assertEqual(columns.website, 12)
        self.assertEqual(columns.website_type, 13)
        self.assertEqual(columns.status, 14)

    def test_existing_result_columns_are_reused_regardless_of_order(self) -> None:
        sheet = Workbook().active
        headers = ["TELEPHONE", "nom_entreprise", "site_web", "siret", "E-MAIL"]
        for column_index, header in enumerate(headers, 1):
            sheet.cell(row=1, column=column_index, value=header)

        columns = resolve_excel_columns(sheet)

        self.assertEqual(columns.company_name, 2)
        self.assertEqual(columns.siret, 4)
        self.assertEqual(columns.email, 5)
        self.assertEqual(columns.phone, 1)
        self.assertEqual(columns.website, 3)
        self.assertIsNone(columns.address)

    def test_official_variant_has_priority_over_input_variant(self) -> None:
        sheet = Workbook().active
        headers = ["siret", "nom_entreprise_input", "nom_entreprise_officiel"]
        for column_index, header in enumerate(headers, 1):
            sheet.cell(row=1, column=column_index, value=header)

        columns = resolve_excel_columns(sheet)

        self.assertEqual(columns.company_name, 3)

    def test_missing_required_header_has_an_actionable_error(self) -> None:
        sheet = Workbook().active
        sheet.cell(row=1, column=1, value="SIRET")

        with self.assertRaisesRegex(MissingExcelColumnError, "company_name"):
            resolve_excel_columns(sheet)

    def test_duplicate_same_header_is_rejected(self) -> None:
        sheet = Workbook().active
        sheet.cell(row=1, column=1, value="SIRET")
        sheet.cell(row=1, column=2, value="siret")
        sheet.cell(row=1, column=3, value="Raison sociale")

        with self.assertRaisesRegex(ValueError, "ambigue"):
            resolve_excel_columns(sheet)


if __name__ == "__main__":
    unittest.main()
