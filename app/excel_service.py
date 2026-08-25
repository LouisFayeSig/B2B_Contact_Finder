from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.excel_columns import ExcelColumns, resolve_excel_columns
from app.models import NOT_FOUND_LABEL, CompanyResult, CompanyRow, ProcessingStatus
from app.utils import clean_cell_value, maybe_limit_rows


class ExcelService:
    def __init__(
        self,
        file_path: Path,
        sheet_name: str | None,
        logger: Any,
        *,
        audit_enabled: bool = False,
    ) -> None:
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.logger = logger
        self.audit_enabled = audit_enabled
        self.columns = ExcelColumns.legacy()
        self._workbook: Workbook | None = None
        self._sheet: Worksheet | None = None

    @property
    def workbook(self) -> Workbook:
        if self._workbook is None:
            raise RuntimeError("Workbook non ouvert.")
        return self._workbook

    @property
    def sheet(self) -> Worksheet:
        if self._sheet is None:
            raise RuntimeError("Worksheet non selectionnee.")
        return self._sheet

    def open(self) -> None:
        self.logger.info("Ouverture du fichier Excel : %s", self.file_path)
        self._workbook = load_workbook(filename=self.file_path)

        if self.sheet_name:
            if self.sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"Feuille introuvable : {self.sheet_name}")
            self._sheet = self.workbook[self.sheet_name]
        else:
            self._sheet = self.workbook.active

        self.columns = resolve_excel_columns(self.sheet)
        self._ensure_result_headers()
        self.logger.info("Feuille selectionnee : %s", self.sheet.title)

    def create_backup(self) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.file_path.with_name(f"{self.file_path.stem}.backup.{timestamp}{self.file_path.suffix}")
        shutil.copy2(self.file_path, backup_path)
        self.logger.info("Backup cree : %s", backup_path)
        return backup_path

    def save(self) -> None:
        self.workbook.save(self.file_path)

    def get_last_useful_row(self) -> int:
        for row_index in range(self.sheet.max_row, 1, -1):
            if any(
                self._read_text(row_index, column_index) is not None
                for column_index in (
                    self.columns.siret,
                    self.columns.company_name,
                    self.columns.address,
                    self.columns.postal_code,
                    self.columns.city,
                )
            ):
                return row_index
        return 1

    def get_row_indexes(self, start_row: int, max_rows: int | None) -> list[int]:
        last_row = self.get_last_useful_row()
        if last_row < start_row:
            return []
        row_indexes = list(range(start_row, last_row + 1))
        return maybe_limit_rows(row_indexes, max_rows)

    def read_company_row(self, row_index: int) -> CompanyRow:
        return CompanyRow(
            row_index=row_index,
            siret=self._read_text(row_index, self.columns.siret) or "",
            company_name=self._read_text(row_index, self.columns.company_name) or "",
            address=self._read_text(row_index, self.columns.address),
            postal_code=self._read_text(row_index, self.columns.postal_code),
            city=self._read_text(row_index, self.columns.city),
        )

    def are_result_cells_filled(self, row_index: int) -> bool:
        values = (
            self._read_text(row_index, self.columns.email),
            self._read_text(row_index, self.columns.phone),
            self._read_text(row_index, self.columns.website),
        )
        return all(value is not None and value != "" for value in values)

    def write_result(
        self,
        row_index: int,
        result: CompanyResult,
        *,
        overwrite_existing: bool = False,
    ) -> None:
        result_values = (
            (self.columns.email, result.email, self.columns.email_source, result.email_source),
            (self.columns.phone, result.phone, self.columns.phone_source, result.phone_source),
            (self.columns.website, result.website, self.columns.website_source, result.website_source),
        )
        website_was_written = False
        website_matches_existing = False
        for column_index, value, source_column, source in result_values:
            existing = self._read_text(row_index, column_index)
            value_was_written = overwrite_existing or not existing or existing == NOT_FOUND_LABEL
            if value_was_written:
                self.sheet.cell(row=row_index, column=column_index, value=value)
            if column_index == self.columns.website:
                website_was_written = value_was_written
                website_matches_existing = self._contact_values_match(column_index, existing, value)
            if self.audit_enabled and (value_was_written or self._contact_values_match(column_index, existing, value)):
                self._write_audit_value(
                    row_index,
                    source_column,
                    source,
                    overwrite_existing=overwrite_existing,
                )

        existing_website_type = self._read_text(row_index, self.columns.website_type)
        if website_was_written or (website_matches_existing and not existing_website_type):
            self.sheet.cell(
                row=row_index,
                column=self.columns.website_type,
                value=result.website_type.value,
            )

        if not self.audit_enabled:
            return

        existing_sources = [] if overwrite_existing else self.read_sources(row_index)
        merged_sources = list(dict.fromkeys([*existing_sources, *result.sources]))[:30]
        if merged_sources or overwrite_existing:
            self.sheet.cell(
                row=row_index,
                column=self.columns.sources,
                value=json.dumps(merged_sources, ensure_ascii=False) if merged_sources else None,
            )

        audit_values: dict[int, str | int] = {
            self.columns.identity_source: result.identity_source,
            self.columns.identity_match_type: result.identity_match_type.value,
            self.columns.searched_at_utc: (
                result.searched_at_utc.isoformat() if result.searched_at_utc is not None else ""
            ),
            self.columns.model_deployment: result.model_deployment,
            self.columns.model_snapshot: result.model_snapshot,
            self.columns.response_id: result.response_id,
            self.columns.input_tokens: result.input_tokens,
            self.columns.output_tokens: result.output_tokens,
            self.columns.total_tokens: result.total_tokens,
            self.columns.web_search_calls: result.web_search_calls,
            self.columns.deterministic_pages: json.dumps(result.deterministic_pages, ensure_ascii=False),
            self.columns.deterministic_fields_found: result.deterministic_fields_found,
            self.columns.email_extraction_method: result.email_extraction_method,
            self.columns.phone_extraction_method: result.phone_extraction_method,
            self.columns.identity_extraction_method: result.identity_extraction_method,
            self.columns.legal_popup_detected: result.legal_popup_detected,
        }
        for column_index, audit_value in audit_values.items():
            self.sheet.cell(
                row=row_index,
                column=column_index,
                value=None if audit_value in ("", NOT_FOUND_LABEL) else audit_value,
            )

    def read_status(self, row_index: int) -> ProcessingStatus | None:
        value = self._read_text(row_index, self.columns.status)
        if value is None:
            return None
        try:
            return ProcessingStatus(value)
        except ValueError:
            return None

    def write_status(self, row_index: int, status: ProcessingStatus) -> None:
        self.sheet.cell(row=row_index, column=self.columns.status, value=status.value)

    def read_sources(self, row_index: int) -> list[str]:
        value = self.sheet.cell(row=row_index, column=self.columns.sources).value
        if not isinstance(value, str) or not value.strip():
            return []
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return [source.strip() for source in value.splitlines() if source.strip()]
        if not isinstance(parsed, list):
            return []
        return [str(source).strip() for source in parsed if str(source).strip()]

    def has_any_contact(self, row_index: int) -> bool:
        values = (
            self._read_text(row_index, self.columns.email),
            self._read_text(row_index, self.columns.phone),
            self._read_text(row_index, self.columns.website),
        )
        return any(value not in (None, "", NOT_FOUND_LABEL) for value in values)

    def read_contact_values(self, row_index: int) -> tuple[str, str, str]:
        return (
            self._read_text(row_index, self.columns.email) or NOT_FOUND_LABEL,
            self._read_text(row_index, self.columns.phone) or NOT_FOUND_LABEL,
            self._read_text(row_index, self.columns.website) or NOT_FOUND_LABEL,
        )

    def _ensure_result_headers(self) -> None:
        headers = {
            self.columns.email: "Email",
            self.columns.phone: "Telephone",
            self.columns.website: "Site Web",
            self.columns.website_type: "Site Web Type",
            self.columns.status: "Enrichment Status",
        }
        if self.audit_enabled:
            headers.update(
                {
                    self.columns.sources: "Enrichment Sources",
                    self.columns.email_source: "Email Source",
                    self.columns.phone_source: "Telephone Source",
                    self.columns.website_source: "Site Web Source",
                    self.columns.identity_source: "Identity Source",
                    self.columns.identity_match_type: "Identity Match Type",
                    self.columns.searched_at_utc: "Search Timestamp UTC",
                    self.columns.model_deployment: "Model Deployment",
                    self.columns.model_snapshot: "Model Snapshot",
                    self.columns.response_id: "Azure Response ID",
                    self.columns.input_tokens: "Input Tokens",
                    self.columns.output_tokens: "Output Tokens",
                    self.columns.total_tokens: "Total Tokens",
                    self.columns.web_search_calls: "Web Search Calls",
                    self.columns.deterministic_pages: "Deterministic Pages",
                    self.columns.deterministic_fields_found: "Deterministic Fields Found",
                    self.columns.email_extraction_method: "Email Extraction Method",
                    self.columns.phone_extraction_method: "Telephone Extraction Method",
                    self.columns.identity_extraction_method: "Identity Extraction Method",
                    self.columns.legal_popup_detected: "Legal Popup Detected",
                }
            )
        for column_index, header in headers.items():
            cell = self.sheet.cell(row=1, column=column_index)
            if cell.value is None:
                cell.value = header

    def _write_audit_value(
        self,
        row_index: int,
        column_index: int,
        value: str,
        *,
        overwrite_existing: bool,
    ) -> None:
        existing = self._read_text(row_index, column_index)
        if overwrite_existing or not existing:
            self.sheet.cell(
                row=row_index,
                column=column_index,
                value=None if value == NOT_FOUND_LABEL else value,
            )

    def _contact_values_match(
        self,
        column_index: int,
        existing: str | None,
        candidate: str,
    ) -> bool:
        if not existing or candidate == NOT_FOUND_LABEL:
            return False
        if column_index == self.columns.email:
            return existing.casefold() == candidate.casefold()
        if column_index == self.columns.phone:
            existing_digits = "".join(character for character in existing if character.isdigit())
            candidate_digits = "".join(character for character in candidate if character.isdigit())
            return existing_digits == candidate_digits
        if column_index == self.columns.website:
            return existing.casefold().rstrip("/") == candidate.casefold().rstrip("/")
        return existing == candidate

    def _read_text(self, row_index: int, column_index: int | None) -> str | None:
        if column_index is None:
            return None
        cell = self.sheet.cell(row=row_index, column=column_index)
        return self._cell_to_text(cell)

    def _cell_to_text(self, cell: Cell) -> str | None:
        value = cell.value
        if value is None:
            return None
        if isinstance(value, str):
            return clean_cell_value(value)
        if isinstance(value, bool):
            return "True" if value else "False"
        if isinstance(value, int):
            return self._format_numeric_cell(cell, value)
        if isinstance(value, float):
            integer_value = int(value) if value.is_integer() else value
            return self._format_numeric_cell(cell, integer_value)
        return clean_cell_value(value)

    def _format_numeric_cell(self, cell: Cell, value: int | float) -> str:
        if isinstance(value, float) and not value.is_integer():
            return str(value)

        integer_value = int(value)
        number_format = str(cell.number_format or "")

        # Preserve les codes postaux stockes comme nombres avec un format du type 00000.
        normalized_format = number_format.replace("\\", "").replace('"', "").strip()
        if "0" in normalized_format and all(char in {"0", "#"} for char in normalized_format):
            width = normalized_format.count("0")
            return str(integer_value).zfill(width)

        return str(integer_value)
