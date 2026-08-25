from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, fields

from openpyxl.worksheet.worksheet import Worksheet


class MissingExcelColumnError(ValueError):
    """Raised when a required business column cannot be identified."""


@dataclass(frozen=True, slots=True)
class ExcelColumns:
    siret: int
    company_name: int
    address: int | None
    postal_code: int | None
    city: int | None
    email: int
    phone: int
    website: int
    status: int
    sources: int
    email_source: int
    phone_source: int
    website_source: int
    identity_source: int
    identity_match_type: int
    searched_at_utc: int
    model_deployment: int
    model_snapshot: int
    response_id: int
    input_tokens: int
    output_tokens: int
    total_tokens: int
    web_search_calls: int
    deterministic_pages: int
    deterministic_fields_found: int
    email_extraction_method: int
    phone_extraction_method: int
    identity_extraction_method: int
    legal_popup_detected: int

    @classmethod
    def legacy(cls) -> ExcelColumns:
        """Keep the historical layout for in-memory use before a workbook is opened."""
        return cls(
            siret=3,
            company_name=4,
            address=7,
            postal_code=8,
            city=9,
            email=16,
            phone=17,
            website=18,
            status=26,
            sources=27,
            email_source=28,
            phone_source=29,
            website_source=30,
            identity_source=31,
            identity_match_type=32,
            searched_at_utc=33,
            model_deployment=34,
            model_snapshot=35,
            response_id=36,
            input_tokens=37,
            output_tokens=38,
            total_tokens=39,
            web_search_calls=40,
            deterministic_pages=41,
            deterministic_fields_found=42,
            email_extraction_method=43,
            phone_extraction_method=44,
            identity_extraction_method=45,
            legal_popup_detected=46,
        )


INPUT_ALIASES: dict[str, tuple[str, ...]] = {
    "siret": ("siret", "numero siret", "no siret", "n siret", "siret etablissement"),
    "company_name": (
        "raison sociale",
        "denomination sociale",
        "denomination",
        "nom entreprise officiel",
        "nom entreprise",
        "nom de l entreprise",
        "nom entreprise input",
        "entreprise",
    ),
    "address": (
        "adresse officielle",
        "adresse",
        "adresse postale",
        "adresse etablissement",
        "adresse input",
    ),
    "postal_code": ("code postal", "codepostal", "cp"),
    "city": ("ville", "commune", "localite"),
}

OUTPUT_ALIASES: dict[str, tuple[str, ...]] = {
    "email": ("email", "e mail", "mail", "adresse email", "email societe"),
    "phone": ("telephone", "tel", "numero telephone"),
    "website": ("site web", "site internet", "website", "url site web"),
    "status": ("enrichment status", "statut enrichissement"),
    "sources": ("enrichment sources", "sources enrichissement"),
    "email_source": ("email source", "source email"),
    "phone_source": ("telephone source", "source telephone"),
    "website_source": ("site web source", "source site web"),
    "identity_source": ("identity source", "source identite"),
    "identity_match_type": ("identity match type", "type rapprochement identite"),
    "searched_at_utc": ("search timestamp utc", "horodatage recherche utc"),
    "model_deployment": ("model deployment", "deploiement modele"),
    "model_snapshot": ("model snapshot", "snapshot modele"),
    "response_id": ("azure response id", "identifiant reponse azure"),
    "input_tokens": ("input tokens", "tokens entree"),
    "output_tokens": ("output tokens", "tokens sortie"),
    "total_tokens": ("total tokens", "tokens total"),
    "web_search_calls": ("web search calls", "appels recherche web"),
    "deterministic_pages": ("deterministic pages", "pages deterministes"),
    "deterministic_fields_found": (
        "deterministic fields found",
        "champs deterministes trouves",
    ),
    "email_extraction_method": ("email extraction method", "methode extraction email"),
    "phone_extraction_method": (
        "telephone extraction method",
        "methode extraction telephone",
    ),
    "identity_extraction_method": (
        "identity extraction method",
        "methode extraction identite",
    ),
    "legal_popup_detected": ("legal popup detected", "popup mentions legales detecte"),
}

REQUIRED_INPUTS = frozenset({"siret", "company_name"})


def normalize_excel_header(value: object) -> str:
    """Return a comparison key insensitive to case, accents and separators."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return " ".join(re.findall(r"[a-z0-9]+", text.casefold()))


def resolve_excel_columns(sheet: Worksheet) -> ExcelColumns:
    """Resolve input and result columns from row-one headers, regardless of order."""
    header_indexes: dict[str, list[int]] = {}
    for header_column_index in range(1, sheet.max_column + 1):
        key = normalize_excel_header(sheet.cell(row=1, column=header_column_index).value)
        if key:
            header_indexes.setdefault(key, []).append(header_column_index)

    resolved: dict[str, int | None] = {}
    missing_required: list[str] = []
    for field_name, aliases in INPUT_ALIASES.items():
        input_column_index = _find_column(header_indexes, aliases, field_name)
        resolved[field_name] = input_column_index
        if input_column_index is None and field_name in REQUIRED_INPUTS:
            missing_required.append(field_name)

    if missing_required:
        expected = ", ".join(missing_required)
        actual = ", ".join(
            str(sheet.cell(row=1, column=index).value)
            for index in range(1, sheet.max_column + 1)
            if sheet.cell(row=1, column=index).value is not None
        )
        raise MissingExcelColumnError(
            f"Colonnes Excel obligatoires introuvables : {expected}. Entetes detectees : {actual or '(aucune)'}."
        )

    next_column = sheet.max_column + 1
    for field_name, aliases in OUTPUT_ALIASES.items():
        output_column_index = _find_column(header_indexes, aliases, field_name)
        if output_column_index is None:
            output_column_index = next_column
            next_column += 1
        resolved[field_name] = output_column_index

    expected_fields = {field.name for field in fields(ExcelColumns)}
    if resolved.keys() != expected_fields:
        raise RuntimeError("La definition des colonnes Excel est incomplete.")
    return ExcelColumns(**resolved)  # type: ignore[arg-type]


def _find_column(
    header_indexes: dict[str, list[int]],
    aliases: tuple[str, ...],
    field_name: str,
) -> int | None:
    for alias in aliases:
        indexes = header_indexes.get(normalize_excel_header(alias), [])
        if len(indexes) > 1:
            raise ValueError(
                f"Entete Excel ambigue pour {field_name} : colonnes {', '.join(map(str, indexes))}."
            )
        if indexes:
            return indexes[0]
    return None
